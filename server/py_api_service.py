from flask import Flask, request, jsonify
import pyodbc
import bcrypt
import os
import re
import unicodedata
import json
import csv
from pathlib import Path

FAILED_DB_SERVERS_LOGGED = set()
EXCEL_CHAT_KB = {
    'name': None,
    'columns': [],
    'rows': [],
}


def escape_html(s):
    return (
        str(s)
        .replace('&', '&amp;')
        .replace('<', '&lt;')
        .replace('>', '&gt;')
        .replace('"', '&quot;')
        .replace("'", '&#39;')
    )


def try_load_json(path: Path):
    try:
        with path.open('r', encoding='utf-8') as f:
            data = json.load(f)
        return data
    except Exception:
        return None


def try_load_csv(path: Path):
    try:
        rows = []
        with path.open('r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            for r in reader:
                rows.append({k: (v if v != '' else None) for k, v in r.items()})
        return {'name': path.name, 'columns': list(rows[0].keys()) if rows else [], 'rows': rows}
    except Exception:
        return None


def try_load_xlsx(path: Path):
    try:
        import openpyxl
    except Exception:
        return None

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        all_rows = []
        all_columns = set()
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            headers = []
            rows = []
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(c) if c is not None else f'col{idx}' for idx, c in enumerate(row)]
                    continue
                obj = {}
                for idx, val in enumerate(row):
                    key = headers[idx] if idx < len(headers) else f'col{idx}'
                    obj[key] = val
                # annotate which sheet this row came from
                obj['Sheet'] = sheet_name
                rows.append(obj)
            for h in headers:
                all_columns.add(h)
        
            all_rows.extend(rows)
        # ensure 'Sheet' is included in columns
        all_columns.add('Sheet')
        return {'name': path.name, 'columns': list(all_columns), 'rows': all_rows}
    except Exception:
        return None


def load_kb_at_startup():
    data_dir = Path(__file__).resolve().parent / 'data'
    data_dir.mkdir(exist_ok=True)

    # Priority: JSON -> CSV -> XLSX
    json_path = data_dir / 'knowledge.json'
    csv_path = data_dir / 'knowledge.csv'
    xlsx_path = data_dir / 'knowledge.xlsx'

    loaded = None
    if json_path.exists():
        loaded = try_load_json(json_path)
        if loaded and isinstance(loaded, dict) and 'rows' in loaded:
            EXCEL_CHAT_KB['name'] = loaded.get('name', json_path.name)
            EXCEL_CHAT_KB['columns'] = loaded.get('columns', [])
            EXCEL_CHAT_KB['rows'] = loaded.get('rows', [])
            print(f"Loaded KB from {json_path} with {len(EXCEL_CHAT_KB['rows'])} rows")
            return

    if csv_path.exists():
        loaded = try_load_csv(csv_path)
        if loaded:
            EXCEL_CHAT_KB['name'] = loaded.get('name')
            EXCEL_CHAT_KB['columns'] = loaded.get('columns', [])
            EXCEL_CHAT_KB['rows'] = loaded.get('rows', [])
            print(f"Loaded KB from {csv_path} with {len(EXCEL_CHAT_KB['rows'])} rows")
            return

    if xlsx_path.exists():
        loaded = try_load_xlsx(xlsx_path)
        if loaded:
            EXCEL_CHAT_KB['name'] = loaded.get('name')
            EXCEL_CHAT_KB['columns'] = loaded.get('columns', [])
            EXCEL_CHAT_KB['rows'] = loaded.get('rows', [])
            print(f"Loaded KB from {xlsx_path} with {len(EXCEL_CHAT_KB['rows'])} rows")
            return

    print('No local KB found at server/data. Start by placing knowledge.json/csv/xlsx into server/data.')


load_kb_at_startup()

app = Flask(__name__)

app = Flask(__name__)


@app.after_request
def add_cors_headers(response):
    configured_origin = os.getenv('CORS_ALLOW_ORIGIN', '*')
    request_origin = request.headers.get('Origin')
    if configured_origin == '*':
        response.headers['Access-Control-Allow-Origin'] = request_origin or '*'
        response.headers['Vary'] = 'Origin'
    else:
        response.headers['Access-Control-Allow-Origin'] = configured_origin
    response.headers['Access-Control-Allow-Credentials'] = 'true'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, PUT, PATCH, DELETE, OPTIONS'
    return response

# --- TICKETS ENDPOINTS ---
@app.route('/tickets', methods=['POST'])
def create_ticket():
    data = request.get_json()
    userId = data.get('userId')
    userEmail = data.get('userEmail')
    title = data.get('title')
    description = data.get('description')
    type_ = data.get('type')
    perimeter = data.get('perimeter')
    format_ = data.get('format')
    urgency = data.get('urgency')
    status = data.get('status', 'En attente')
    if not all([title, description, type_, perimeter, format_, urgency]):
        return jsonify({'error': 'Tous les champs obligatoires doivent être remplis'}), 400
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute(
        '''
        INSERT INTO tickets (userId, userEmail, title, description, type, perimeter, format, urgency, status, createdAt)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, GETDATE())
        ''',
        (userId, userEmail, title, description, type_, perimeter, format_, urgency, status)
    )
    conn.commit()
    conn.close()
    return jsonify({'message': 'Ticket created successfully'}), 201

@app.route('/tickets', methods=['GET'])
def get_tickets():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('''SELECT id, userId, userEmail, title, description, type, perimeter, format, urgency, status, createdAt, updatedAt, solution, solvedBy, solvedAt FROM tickets ORDER BY createdAt DESC''')
    tickets = [
        dict(zip([column[0] for column in cursor.description], row))
        for row in cursor.fetchall()
    ]
    conn.close()
    return jsonify(tickets)

@app.route('/tickets/<int:ticket_id>', methods=['PATCH'])
def update_ticket(ticket_id):
    data = request.get_json()
    status = data.get('status')
    solution = data.get('solution')
    solvedBy = data.get('solvedBy')
    update_fields = []
    update_values = []
    if status:
        update_fields.append('status=?')
        update_values.append(status)
    if solution:
        update_fields.append('solution=?')
        update_values.append(solution)
    if solvedBy:
        update_fields.append('solvedBy=?')
        update_values.append(solvedBy)
        update_fields.append('solvedAt=GETDATE()')
    update_fields.append('updatedAt=GETDATE()')
    if not update_fields:
        return jsonify({'error': 'No fields to update'}), 400
    sql = f"UPDATE tickets SET {', '.join(update_fields)} WHERE id=?"
    update_values.append(ticket_id)
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute(sql, tuple(update_values))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Ticket updated'}), 200


# --- REPORTS ENDPOINTS ---
@app.route('/reports', methods=['POST'])
def create_report():
    data = request.get_json()
    userId = data.get('userId')
    userEmail = data.get('userEmail')
    title = data.get('title')
    description = data.get('description')
    if not title or not description:
        return jsonify({'error': 'Title and description required'}), 400
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute(
        '''
        INSERT INTO reports (userId, userEmail, title, description, status, createdAt)
        VALUES (?, ?, ?, ?, 'open', GETDATE())
        ''',
        (userId, userEmail, title, description)
    )
    conn.commit()
    conn.close()
    return jsonify({'message': 'Report created successfully'}), 201

# List all reports
@app.route('/reports', methods=['GET'])
def get_reports():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('''SELECT id, userId, userEmail, title, description, status, createdAt, updatedAt, solvedBy, solvedAt, solution FROM reports ORDER BY createdAt DESC''')
    reports = [
        dict(zip([column[0] for column in cursor.description], row))
        for row in cursor.fetchall()
    ]
    conn.close()
    return jsonify(reports)

# Solve a report (admin)
@app.route('/reports/<int:report_id>', methods=['PATCH'])
def solve_report(report_id):
    data = request.get_json()
    solution = data.get('solution')
    solvedBy = data.get('solvedBy')
    if not solution or not solvedBy:
        return jsonify({'error': 'Solution and solvedBy required'}), 400
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute(
        '''
        UPDATE reports SET status='solved', solution=?, solvedBy=?, solvedAt=GETDATE(), updatedAt=GETDATE() WHERE id=?
        ''',
        (solution, solvedBy, report_id)
    )
    conn.commit()
    conn.close()
    return jsonify({'message': 'Report solved'}), 200


# --- LOGS ENDPOINTS ---
@app.route('/logs', methods=['POST'])
def add_log():
    data = request.get_json()
    userId = data.get('userId')
    userEmail = data.get('userEmail')
    action = data.get('action')
    details = data.get('details', '')
    ip = data.get('ip')
    if not action:
        return jsonify({'error': 'Action is required'}), 400
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute(
        """
        INSERT INTO logs (userId, userEmail, action, details, ip, createdAt)
        VALUES (?, ?, ?, ?, ?, GETDATE())
        """,
        (userId, userEmail, action, details, ip)
    )
    conn.commit()
    conn.close()
    return jsonify({'message': 'Log added successfully'}), 201

@app.route('/logs', methods=['GET'])
def get_logs():
    limit = int(request.args.get('limit', 100))
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute(f"SELECT TOP {limit} id, userId, userEmail, action, details, ip, createdAt FROM logs ORDER BY createdAt DESC")
    logs = [
        dict(zip([column[0] for column in cursor.description], row))
        for row in cursor.fetchall()
    ]
    conn.close()
    return jsonify(logs)


@app.route('/users', methods=['POST'])
def create_user():
    data = request.get_json()
    email = data.get('email')
    password = data.get('password')
    fullName = data.get('fullName', email)
    role = data.get('role', 'user')
    isActive = data.get('isActive', True)

    if not email or not password:
        return jsonify({'error': 'Email and password required'}), 400

    password_hash = bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO users (email, passwordHash, fullName, role, isActive) VALUES (?, ?, ?, ?, ?)",
        (email, password_hash, fullName, role, isActive)
    )
    conn.commit()
    conn.close()
    return jsonify({'message': 'User created successfully', 'email': email}), 201

@app.before_request
def log_request_info():
    print('Request:', request.method, request.path)

def get_db_connection():
    def normalize_server_name(raw_server):
        if not raw_server:
            return ""
        current = str(raw_server).strip().strip('"').strip("'")
        current = current.replace('\\\\', '\\')
        if current.lower().startswith('np:') and '\\\\' in current:
            current = current.replace('\\\\', '\\')
        return current

    database = os.getenv('DB_DATABASE', 'DataFlow')
    timeout = int(os.getenv('DB_CONNECT_TIMEOUT', '5'))
    candidate_servers = []

    env_server = os.getenv('DB_SERVER', r'(localdb)\MyInstance')
    if env_server:
        candidate_servers.append(env_server)

    env_named_pipe = os.getenv('DB_NAMED_PIPE')
    if env_named_pipe:
        candidate_servers.append(env_named_pipe)

    # Try common SQL Server instances
    for fallback_server in [
        r'(localdb)\MyInstance',
        r'(localdb)\MSSQLLocalDB',
        r'localhost\SQLEXPRESS',
        r'.\SQLEXPRESS',
        r'127.0.0.1\SQLEXPRESS',
        'localhost',
        '.',
    ]:
        if fallback_server not in candidate_servers:
            candidate_servers.append(fallback_server)

    normalized_candidates = []
    for server in candidate_servers:
        current = normalize_server_name(server)
        if current and current not in normalized_candidates:
            normalized_candidates.append(current)

    last_error = None
    for server in normalized_candidates:
        conn_str = (
            r'DRIVER={ODBC Driver 17 for SQL Server};'
            f'SERVER={server};'
            f'DATABASE={database};'
            r'Trusted_Connection=yes;'
            r'TrustServerCertificate=yes;'
        )
        try:
            return pyodbc.connect(conn_str, timeout=timeout)
        except Exception as e:
            last_error = e
            if server not in FAILED_DB_SERVERS_LOGGED:
                print(f"[DB CONNECTION TRY FAILED] server={server} error={e}")
                FAILED_DB_SERVERS_LOGGED.add(server)

    print(f"[DB CONNECTION ERROR] Could not connect using servers={normalized_candidates}. Last error: {last_error}")
    raise last_error

@app.route('/login', methods=['POST'])
def login():
    data = request.get_json()
    email = data.get('email')
    password = data.get('password')
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT id, email, fullName, role, isActive, passwordHash FROM users WHERE email = ?", email)
    row = cursor.fetchone()
    conn.close()
    if not row:
        return jsonify({'error': 'Invalid credentials'}), 401
    user = {
        'id': row[0],
        'email': row[1],
        'fullName': row[2],
        'role': row[3],
        'isActive': row[4],
    }
    password_hash = row[5]
    if not bcrypt.checkpw(password.encode(), password_hash.encode()):
        return jsonify({'error': 'Invalid credentials'}), 401
    return jsonify({'message': 'Login successful', 'user': user})


def normalize_text(value):
    text = unicodedata.normalize('NFD', str(value or ''))
    text = ''.join(character for character in text if unicodedata.category(character) != 'Mn')
    return text.lower()


def tokenize(value):
    raw_tokens = re.findall(r'[A-Za-z0-9]{2,}', normalize_text(value))
    stop_words = {
        'ai', 'alors', 'au', 'aux', 'avec', 'ce', 'ces', 'cette', 'dans', 'de', 'des', 'du',
        'elle', 'en', 'est', 'et', 'fait', 'faire', 'il', 'je', 'la', 'le', 'les', 'leur',
        'leurs', 'lui', 'mais', 'me', 'mes', 'mon', 'ne', 'nos', 'notre', 'nous', 'on', 'ou',
        'par', 'pas', 'pour', 'que', 'qui', 'sa', 'se', 'ses', 'son', 'sur', 'ta', 'te', 'tes',
        'ton', 'tu', 'un', 'une', 'vos', 'votre', 'vous'
    }
    return [token for token in raw_tokens if token not in stop_words]


def row_to_search_text(row):
    return ' | '.join(f'{key}: {value}' for key, value in row.items())


def score_row(row, tokens):
    if not tokens:
        return 0
    haystack = normalize_text(row_to_search_text(row))
    return sum(1 for token in tokens if token in haystack)


def format_row(row):
    parts = []
    for key, value in list(row.items())[:8]:
        if value is None:
            continue
        text = str(value).strip()
        if text:
            parts.append(f'{key}: {text}')
    return ' | '.join(parts) if parts else 'Ligne vide'


def bold_icd_codes(text):
    # Return ICD codes as plain text; normalize slashes and spacing
    escaped = escape_html(text)
    return re.sub(r'\s*/\s*', ' / ', escaped)


def build_excel_reply(question, mode='synthese'):
    if not EXCEL_CHAT_KB['rows']:
        return "Aucun fichier Excel n'est chargé. Importez d'abord votre Excel pour que je réponde uniquement à partir de ce contenu."

    question_upper = str(question or '').upper()
    exact_codes = re.findall(r'[A-Z]\d{2,4}', question_upper)
    tokens = tokenize(question)
    ranked_rows = []
    for index, row in enumerate(EXCEL_CHAT_KB['rows']):
        score = score_row(row, tokens)
        row_upper = normalize_text(row_to_search_text(row)).upper()
        if exact_codes and any(code in row_upper for code in exact_codes):
            score += 5
        # If row has a category-like column name, boost score when tokens match it
        if tokens:
            for key, val in row.items():
                if val is None:
                    continue
                key_norm = normalize_text(str(key))
                if any(k in key_norm for k in ['categorie', 'category', 'cat', 'specialite']):
                    if any(token in normalize_text(str(val)) for token in tokens):
                        score += 4
        if score > 0:
            ranked_rows.append((score, index, row))
    ranked_rows.sort(key=lambda item: (-item[0], item[1]))

    def extract_category_reference(row):
        def pick_value_priority(keywords):
            # For priority, iterate keywords in order and for each keyword
            # scan all keys to find the first matching key.
            for keyword in keywords:
                for key, value in row.items():
                    if value is None:
                        continue
                    key_norm = normalize_text(str(key))
                    if keyword in key_norm:
                        text = str(value).strip()
                        if text:
                            return text
            return None

        category = pick_value_priority(['categorie', 'category', 'cat', 'specialite'])
        # prefer ICD/code columns before acte/soin
        reference = pick_value_priority(['icd', 'code', 'reference', 'ref', 'acte', 'soin'])

        if category and reference:
            return category, reference

        # Fallback: first two non-empty values if headers are unknown.
        non_empty_values = []
        for _, value in row.items():
            if value is None:
                continue
            text = str(value).strip()
            if text:
                non_empty_values.append(text)

        if not category and non_empty_values:
            category = non_empty_values[0]
        if not reference and len(non_empty_values) > 1:
            reference = non_empty_values[1]

        return category, reference

    def find_effective_category_at(index):
        # Look at current and previous rows to find nearest non-empty category-like field
        for j in range(index, -1, -1):
            r = EXCEL_CHAT_KB['rows'][j]
            cat, _ = extract_category_reference(r)
            if cat:
                return cat
        return None

    if not ranked_rows:
        # Fallback: try to match the query against obvious category/speciality columns
        qnorm = normalize_text(str(question or ''))
        candidate_rows = []
        for index, row in enumerate(EXCEL_CHAT_KB['rows']):
            # try to extract category-like fields, fallback to previous rows if empty
            category, _ = extract_category_reference(row)
            if not category:
                category = find_effective_category_at(index)
            if category:
                if normalize_text(category).find(qnorm) != -1:
                    candidate_rows.append((0, index, row))

        if candidate_rows:
            # format candidate rows similarly to synthesis mode
            parts = []
            for _, _, row in candidate_rows[:6]:
                cat, ref = extract_category_reference(row)
                frags = []
                if cat:
                    frags.append(f"Catégorie: {escape_html(cat)}")
                if ref:
                    frags.append(f"Référence: {bold_icd_codes(ref)}")
                parts.append(' — '.join(frags))
            return '<br/>'.join(parts)

        # Second fallback: search any cell content (soins, ICD, etc.) for query substring
        cell_matches = []
        for index, row in enumerate(EXCEL_CHAT_KB['rows']):
            # join all values into a single searchable string
            row_values = ' '.join([str(v) for v in row.values() if v is not None])
            if normalize_text(row_values).find(qnorm) != -1:
                cell_matches.append((0, index, row))

        if cell_matches:
            parts = []
            for _, _, row in cell_matches[:6]:
                cat = find_effective_category_at(EXCEL_CHAT_KB['rows'].index(row))
                _, ref = extract_category_reference(row)
                frags = []
                if cat:
                    frags.append(f"Catégorie: {escape_html(cat)}")
                if ref:
                    frags.append(f"Référence: {bold_icd_codes(ref)}")
                # include a short snippet of the soins cell
                soins = row.get('Soins') or row.get('Soin') or None
                if soins:
                    frags.append(f"Soin: {escape_html(str(soins))}")
                parts.append(' — '.join(frags))
            return '<br/>'.join(parts)

        return "Je ne trouve pas cette information. Posez une question plus proche des colonnes ou des valeurs du tableau."

    def highlight(text, question):
        # Return escaped plain text (no HTML tags)
        return escape_html(text)

    def extract_category_reference(row):
        def pick_value_priority(keywords):
            for keyword in keywords:
                for key, value in row.items():
                    if value is None:
                        continue
                    key_norm = normalize_text(str(key))
                    if keyword in key_norm:
                        text = str(value).strip()
                        if text:
                            return text
            return None

        category = pick_value_priority(['categorie', 'category', 'cat', 'specialite'])
        reference = pick_value_priority(['icd', 'code', 'reference', 'ref', 'acte', 'soin'])

        if category and reference:
            return category, reference

        # Fallback: first two non-empty values if headers are unknown.
        non_empty_values = []
        for _, value in row.items():
            if value is None:
                continue
            text = str(value).strip()
            if text:
                non_empty_values.append(text)

        if not category and non_empty_values:
            category = non_empty_values[0]
        if not reference and len(non_empty_values) > 1:
            reference = non_empty_values[1]

        return category, reference

    # STRICT MODE: return only the matching rows (no extra synthesis)
    if mode == 'strict':
        lines = []
        for _, _, row in ranked_rows[:3]:
            lines.append(highlight(format_row(row), question))
        return '<br/>'.join(lines)

    # SYNTHESIS MODE: return only category/reference pairs for matched rows.
    # If the user's query is exactly a category, return only the unique references for that category
    qnorm = normalize_text(str(question or ''))
    # Build mapping category -> list of rows
    category_map = {}
    for idx, r in enumerate(EXCEL_CHAT_KB['rows']):
        cat, ref = extract_category_reference(r)
        if not cat:
            cat = find_effective_category_at(idx)
        if cat:
            key = normalize_text(cat)
            category_map.setdefault(key, []).append((cat, ref, r))

    # exact category match first
    if qnorm in category_map:
        entries = category_map[qnorm]
        seen = set()
        parts = []
        for cat, ref, _ in entries:
            key = (normalize_text(cat), normalize_text(str(ref or '')))
            if key in seen:
                continue
            seen.add(key)
            frags = []
            if cat:
                frags.append(f"Catégorie: {escape_html(cat)}")
            if ref:
                frags.append(f"Référence: {escape_html(ref)}")
            parts.append(' — '.join(frags))
        return '<br/>'.join(parts) if parts else "Je ne trouve pas de référence pour cette catégorie."

    # SYNTHESIS MODE: return only category/reference pairs for matched rows (deduplicated)
    summary_parts = []
    seen = set()
    for _, (_, _, row) in enumerate(ranked_rows[:6]):
        category, reference = extract_category_reference(row)
        key = (normalize_text(str(category or '')), normalize_text(str(reference or '')))
        if key in seen:
            continue
        seen.add(key)
        fragments = []
        if category:
            fragments.append(f"Catégorie: {escape_html(category)}")
        if reference:
            fragments.append(f"Référence: {escape_html(reference)}")
        row_text = ' — '.join(fragments) if fragments else highlight(format_row(row), question)
        row_text = highlight(row_text, question)
        summary_parts.append(row_text)

    return '<br/>'.join(summary_parts) if summary_parts else "Je ne trouve pas cette information."


@app.route('/api/chat/knowledge-base', methods=['POST'])
def upload_chat_knowledge_base():
    data = request.get_json() or {}
    rows = data.get('rows')

    if not isinstance(rows, list):
        return jsonify({'error': 'Le champ "rows" doit être un tableau.'}), 400

    normalized_rows = []
    for row in rows:
        if isinstance(row, dict):
            normalized_row = {str(key): value for key, value in row.items()}
            normalized_rows.append(normalized_row)

    EXCEL_CHAT_KB['name'] = data.get('name').strip() if isinstance(data.get('name'), str) and data.get('name').strip() else 'Excel importé'
    EXCEL_CHAT_KB['columns'] = [str(column) for column in data.get('columns', []) if column]
    EXCEL_CHAT_KB['rows'] = normalized_rows

    # Persist KB to server/data/knowledge.json so it's preloaded on restart
    try:
        data_dir = Path(__file__).resolve().parent / 'data'
        data_dir.mkdir(exist_ok=True)
        out_path = data_dir / 'knowledge.json'
        with out_path.open('w', encoding='utf-8') as f:
            json.dump({
                'name': EXCEL_CHAT_KB['name'],
                'columns': EXCEL_CHAT_KB['columns'],
                'rows': EXCEL_CHAT_KB['rows'],
            }, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"[KB PERSIST ERROR] {e}")

    return jsonify({
        'message': 'Base Excel chargée',
        'name': EXCEL_CHAT_KB['name'],
        'rowCount': len(EXCEL_CHAT_KB['rows']),
        'columns': EXCEL_CHAT_KB['columns'],
    })


@app.route('/api/chat/knowledge-base', methods=['GET'])
def knowledge_base_status():
    loaded = bool(EXCEL_CHAT_KB.get('rows'))
    return jsonify({
        'loaded': loaded,
        'name': EXCEL_CHAT_KB.get('name'),
        'rowCount': len(EXCEL_CHAT_KB.get('rows', [])),
        'columns': EXCEL_CHAT_KB.get('columns', []),
    })


@app.route('/api/chat/knowledge-base/categories', methods=['GET'])
def knowledge_base_categories():
    # Return unique effective categories from the loaded KB (preserve order)
    seen = set()
    categories = []
    rows = EXCEL_CHAT_KB.get('rows', [])
    def extract_cat_at(i):
        # find effective category for row i
        for j in range(i, -1, -1):
            r = rows[j]
            # similar logic to extract_category_reference
            for key, value in r.items():
                if value is None:
                    continue
                key_norm = normalize_text(str(key))
                if any(k in key_norm for k in ['categorie', 'category', 'cat', 'specialite']):
                    text = str(value).strip()
                    if text:
                        return text
        return None

    for idx, r in enumerate(rows):
        cat = extract_cat_at(idx)
        if cat and cat not in seen:
            seen.add(cat)
            categories.append(cat)

    return jsonify({'categories': categories})


@app.route('/api/chat', methods=['POST'])
def chat_reply():
    data = request.get_json() or {}
    messages = data.get('messages')

    if not isinstance(messages, list):
        return jsonify({'error': 'Le champ "messages" est requis et doit être un tableau.'}), 400

    latest_user_message = None
    for message in reversed(messages):
        if isinstance(message, dict) and message.get('role') == 'user':
            latest_user_message = message.get('content', '')
            break

    if not str(latest_user_message).strip():
        return jsonify({'error': 'Un message utilisateur est requis.'}), 400

    mode = data.get('mode', 'synthese')
    return jsonify({'reply': build_excel_reply(str(latest_user_message), mode=mode)})

@app.route('/users', methods=['GET'])
def get_users():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT id, email, fullName, role, isActive FROM users")
    users = [
        dict(zip([column[0] for column in cursor.description], row))
        for row in cursor.fetchall()
    ]
    conn.close()
    return jsonify(users)

# ============== BATCH PROCESSING ENDPOINTS ==============
@app.route('/api/batch', methods=['GET'])
def get_batches():
    """Get all batch processes from database"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT id, nom, description, categorie, parametres, actif FROM BatchProcesses ORDER BY nom")
        batches = [
            dict(zip([column[0] for column in cursor.description], row))
            for row in cursor.fetchall()
        ]
        conn.close()
        return jsonify(batches)
    except Exception as e:
        print(f"[BATCH ERROR] {e}")
        # Return empty list if table doesn't exist yet
        return jsonify([])

@app.route('/api/batch/<int:batch_id>/execute', methods=['POST'])
def execute_batch(batch_id):
    """Execute a batch process with parameters"""
    try:
        data = request.get_json()
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Get batch definition
        cursor.execute("SELECT nom, requete_sql, parametres FROM BatchProcesses WHERE id = ?", batch_id)
        batch = cursor.fetchone()
        
        if not batch:
            return jsonify({'error': 'Batch not found'}), 404
        
        batch_name, sql_query, params_str = batch[0], batch[1], batch[2]
        
        # Replace parameters in SQL query
        if data:
            for key, value in data.items():
                if value:
                    sql_query = sql_query.replace(f"{{{key}}}", f"'{value}'")
        
        # Execute the query
        cursor.execute(sql_query)
        rows = cursor.fetchall()
        result = [
            dict(zip([column[0] for column in cursor.description], row))
            for row in rows
        ]
        conn.close()
        
        return jsonify({
            'batch': batch_name,
            'total': len(result),
            'data': result
        })
    except Exception as e:
        print(f"[BATCH EXECUTE ERROR] {e}")
        return jsonify({'error': str(e)}), 500

# ============== DASHBOARD DATA ENDPOINTS ==============
def _rows_to_dicts(cursor):
    columns = [column[0] for column in cursor.description]
    return [dict(zip(columns, row)) for row in cursor.fetchall()]


def _build_claim_filters(args):
    clauses = []
    params = []

    centre = args.get('centre')
    centres = args.get('centres')
    date_debut = args.get('date_debut') or args.get('dateDebut')
    date_fin = args.get('date_fin') or args.get('dateFin')
    marque = args.get('marque')
    nom_assureur = args.get('nom_assureur') or args.get('nomAssureur')
    team = args.get('team')
    agent = args.get('agent') or args.get('nameAgent')

    if centre:
        clauses.append("site_gestion_theo = ?")
        params.append(centre)
    elif centres:
        values = [value.strip() for value in centres.split(',') if value.strip()]
        if values:
            placeholders = ','.join(['?'] * len(values))
            clauses.append(f"site_gestion_theo IN ({placeholders})")
            params.extend(values)
    if date_debut:
        clauses.append("CAST(date_survenance AS DATE) >= CAST(? AS DATE)")
        params.append(date_debut)
    if date_fin:
        clauses.append("CAST(date_survenance AS DATE) <= CAST(? AS DATE)")
        params.append(date_fin)
    if marque:
        clauses.append("marque = ?")
        params.append(marque)
    if nom_assureur:
        clauses.append("nom_assureur = ?")
        params.append(nom_assureur)
    if team:
        clauses.append("site_gestion_theo = ?")
        params.append(team)
    if agent:
        clauses.append("utilisateur = ?")
        params.append(agent)

    where_sql = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    return where_sql, params


CLAIM_COLUMN_CACHE = {}


def _has_column(conn, table_name, column_name):
    cache_key = f"{table_name}.{column_name}"
    if cache_key in CLAIM_COLUMN_CACHE:
        return CLAIM_COLUMN_CACHE[cache_key]

    cursor = conn.cursor()
    cursor.execute(
        """
        SELECT 1
        FROM INFORMATION_SCHEMA.COLUMNS
        WHERE TABLE_SCHEMA = 'dbo'
          AND TABLE_NAME = ?
          AND COLUMN_NAME = ?
        """,
        (table_name, column_name),
    )
    exists = cursor.fetchone() is not None
    CLAIM_COLUMN_CACHE[cache_key] = exists
    return exists


def _rejection_reason_expression(has_column):
    if has_column:
        return "NULLIF(LTRIM(RTRIM(c.rejection_reason)), '')"

    return "CASE WHEN UPPER(ISNULL(c.code_etat, '')) IN ('REJETE', 'RJ', 'REJECTED') THEN 'Rejeté' ELSE NULL END"


@app.route('/api/data/filters', methods=['GET'])
def get_filters():
    """Get available filters for dashboard from real claims table."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("SELECT DISTINCT marque FROM dbo.claims_closed WHERE marque IS NOT NULL AND LTRIM(RTRIM(marque)) <> '' ORDER BY marque")
        marques = [row[0] for row in cursor.fetchall()]

        cursor.execute("SELECT DISTINCT nom_assureur FROM dbo.claims_closed WHERE nom_assureur IS NOT NULL AND LTRIM(RTRIM(nom_assureur)) <> '' ORDER BY nom_assureur")
        assureurs = [row[0] for row in cursor.fetchall()]

        cursor.execute("SELECT DISTINCT site_gestion_theo FROM dbo.claims_closed WHERE site_gestion_theo IS NOT NULL AND LTRIM(RTRIM(site_gestion_theo)) <> '' ORDER BY site_gestion_theo")
        teams = [row[0] for row in cursor.fetchall()]

        cursor.execute("SELECT DISTINCT utilisateur FROM dbo.claims_closed WHERE utilisateur IS NOT NULL AND LTRIM(RTRIM(utilisateur)) <> '' ORDER BY utilisateur")
        agents = [row[0] for row in cursor.fetchall()]

        conn.close()
        return jsonify({
            'marques': marques,
            'assureurs': assureurs,
            'teams': teams,
            'agents': agents
        })
    except Exception as e:
        print(f"[FILTERS ERROR] {e}")
        return jsonify({'marques': [], 'assureurs': [], 'teams': [], 'agents': []})


@app.route('/api/data/dashboard', methods=['GET'])
def get_dashboard_data():
    """Get dashboard KPI data from correct line tables."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # Count terminated lines
        cursor.execute("SELECT COUNT(*) AS termines FROM dbo.sinistres_termines_ligne")
        term_row = cursor.fetchone()
        termines = int(term_row[0] or 0)

        # Count non-terminated lines
        cursor.execute("SELECT COUNT(*) AS en_cours FROM dbo.sinistre_Ter_Lignes_NO_TE")
        nt_row = cursor.fetchone()
        en_cours = int(nt_row[0] or 0)

        total = termines + en_cours

        conn.close()

        return jsonify({
            'sinistres': {
                'total': total,
                'termines': termines,
                'en_cours': en_cours,
                'rejetes': 0,
                'total_soumis': 0,
                'total_rembourse': 0
            },
            'parCentre': []
        })
    except Exception as e:
        print(f"[DASHBOARD ERROR] {e}")
        return jsonify({'sinistres': {}, 'parCentre': []})


@app.route('/api/data/teams', methods=['GET'])
def get_teams():
    """Get team performance data from real claims table."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        where_sql, params = _build_claim_filters(request.args)

        cursor.execute(
            f"""
            SELECT
                ISNULL(c.site_gestion_theo, 'N/A') AS team,
                ISNULL(c.utilisateur, 'N/A') AS agent,
                ISNULL(c.site_gestion_theo, 'N/A') AS centre,
                CAST(AVG(CASE WHEN c.date_cloture IS NOT NULL AND c.date_survenance IS NOT NULL AND DATEDIFF(DAY, c.date_survenance, c.date_cloture) <= 20 THEN 100.0 ELSE 0.0 END) AS DECIMAL(10,2)) AS provider_claim_20days,
                SUM(CASE WHEN c.date_cloture IS NOT NULL AND c.date_survenance IS NOT NULL AND DATEDIFF(DAY, c.date_survenance, c.date_cloture) <= 20 THEN 1 ELSE 0 END) AS nb_claim_treated_20d,
                SUM(CASE WHEN c.date_cloture IS NOT NULL AND c.date_survenance IS NOT NULL AND DATEDIFF(DAY, c.date_survenance, c.date_cloture) <= 30 THEN 1 ELSE 0 END) AS nb_claim_treated_30d,
                COUNT(DISTINCT c.id_sinistre) AS nb_claims_treated,
                COUNT(l.id_ligne) AS nb_ligne_claims_treated,
                COUNT(DISTINCT CASE WHEN l.id_ligne IS NOT NULL THEN c.id_sinistre END) AS nb_claims_ss,
                COUNT(l.id_ligne) AS nb_ligne_claims_ss
                        FROM dbo.claims_closed c
            LEFT JOIN dbo.claims_closed_lines l
              ON l.id_sinistre = c.id_sinistre AND l.num_sinistre = c.num_sinistre
            {where_sql}
            GROUP BY c.site_gestion_theo, c.utilisateur
            ORDER BY nb_claims_treated DESC, team ASC, agent ASC
            """,
            params
        )
        teams = _rows_to_dicts(cursor)
        conn.close()
        return jsonify(teams)
    except Exception as e:
        print(f"[TEAMS ERROR] {e}")
        return jsonify([])


@app.route('/api/data/sinistres', methods=['GET'])
def get_sinistres():
    """Get non-terminated lines from correct table."""
    import sys
    import traceback
    logfile = open('C:\\Work\\DataFlow\\server\\debug_sinistres.log', 'a')
    
    logfile.write("[SINISTRES_START] === SINISTRES REQUEST ===\n")
    logfile.flush()
    try:
        logfile.write("[SINISTRES] Getting DB connection...\n")
        logfile.flush()
        conn = get_db_connection()
        logfile.write(f"[SINISTRES] Connection type: {type(conn)}\n")
        logfile.flush()
        logfile.write("[SINISTRES] Creating cursor...\n")
        logfile.flush()
        cursor = conn.cursor()
        logfile.write("[SINISTRES] Executing query...\n")
        logfile.flush()
        cursor.execute("SELECT TOP 500 * FROM dbo.sinistre_Ter_Lignes_NO_TE ORDER BY ID_Sinistre DESC")
        logfile.write("[SINISTRES] Fetching all results...\n")
        logfile.flush()
        sinistres = _rows_to_dicts(cursor)
        logfile.write(f"[SINISTRES] SUCCESS! Got {len(sinistres)} rows\n")
        logfile.flush()
        conn.close()
        logfile.close()
        return jsonify(sinistres)
    except Exception as e:
        logfile.write(f"[SINISTRES ERROR] Exception occurred: {e}\n")
        logfile.write(f"[SINISTRES ERROR] Full traceback:\n")
        traceback.print_exc(file=logfile)
        logfile.flush()
        logfile.close()
        return jsonify([])


@app.route('/api/data/sinistres-termines', methods=['GET'])
def get_sinistres_termines():
    """Get terminated lines from correct table."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT TOP 1000 * FROM dbo.sinistres_termines_ligne ORDER BY ID_Sinistre DESC")
        sinistres = _rows_to_dicts(cursor)
        conn.close()
        return jsonify(sinistres)
    except Exception as e:
        print(f"[SINISTRES TERMINES ERROR] {e}")
        return jsonify([])


@app.route('/api/data/monthly-trends', methods=['GET'])
def get_monthly_trends():
    """Get monthly claim trends from real claims table."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT
                YEAR(date_cloture) AS annee,
                MONTH(date_cloture) AS mois_num,
                COUNT(*) AS sinistres
            FROM dbo.claims_closed
            WHERE date_cloture IS NOT NULL
            GROUP BY YEAR(date_cloture), MONTH(date_cloture)
            ORDER BY annee, mois_num
            """
        )

        month_labels = {
            1: 'Jan', 2: 'Fév', 3: 'Mar', 4: 'Avr', 5: 'Mai', 6: 'Juin',
            7: 'Juil', 8: 'Aoû', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Déc'
        }
        trends = []
        for row in cursor.fetchall():
            mois_num = int(row[1])
            trends.append({
                'mois': month_labels.get(mois_num, str(mois_num)),
                'sinistres': int(row[2] or 0)
            })

        conn.close()
        return jsonify(trends)
    except Exception as e:
        print(f"[MONTHLY TRENDS ERROR] {e}")
        return jsonify([])


@app.route('/api/data/top-ec-reasons', methods=['GET'])
def get_top_ec_reasons():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        where_sql, params = _build_claim_filters(request.args)
        has_ec_reason = _has_column(conn, 'claims_closed', 'EC_REASON')
        reason_expr = "NULLIF(LTRIM(RTRIM(EC_REASON)), '')" if has_ec_reason else "NULLIF(LTRIM(RTRIM(type_of_log)), '')"

        cursor.execute(
            f"""
            SELECT TOP 10
                COALESCE({reason_expr}, 'Autre') AS name,
                COUNT(*) AS value
            FROM dbo.claims_closed
            {where_sql}
            GROUP BY COALESCE({reason_expr}, 'Autre')
            ORDER BY value DESC, name ASC
            """,
            params,
        )
        rows = _rows_to_dicts(cursor)
        conn.close()
        return jsonify(rows)
    except Exception as e:
        print(f"[TOP EC REASONS ERROR] {e}")
        return jsonify([])


@app.route('/api/data/adjustment-reasons', methods=['GET'])
def get_adjustment_reasons():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        where_sql, params = _build_claim_filters(request.args)
        has_adjustment_reason = _has_column(conn, 'claims_closed', 'ADJUSTMENT_REASON')
        reason_expr = "NULLIF(LTRIM(RTRIM(ADJUSTMENT_REASON)), '')" if has_adjustment_reason else "NULLIF(LTRIM(RTRIM(type_of_log)), '')"

        cursor.execute(
            f"""
            SELECT TOP 10
                COALESCE({reason_expr}, 'Autre') AS name,
                COUNT(*) AS value
            FROM dbo.claims_closed
            {where_sql}
            GROUP BY COALESCE({reason_expr}, 'Autre')
            ORDER BY value DESC, name ASC
            """,
            params,
        )
        rows = _rows_to_dicts(cursor)
        conn.close()
        return jsonify(rows)
    except Exception as e:
        print(f"[ADJUSTMENT REASONS ERROR] {e}")
        return jsonify([])


@app.route('/api/data/top-rejection-reasons', methods=['GET'])
def get_top_rejection_reasons():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        where_sql, params = _build_claim_filters(request.args)
        has_rejection_reason = _has_column(conn, 'claims_closed', 'rejection_reason')
        rejection_expr = _rejection_reason_expression(has_rejection_reason)

        cursor.execute(
            f"""
            SELECT TOP 10
                COALESCE({rejection_expr}, 'Autre') AS name,
                COUNT(*) AS value
            FROM dbo.claims_closed c
            {where_sql}
            GROUP BY COALESCE({rejection_expr}, 'Autre')
            ORDER BY value DESC, name ASC
            """,
            params,
        )
        rows = _rows_to_dicts(cursor)
        conn.close()
        return jsonify(rows)
    except Exception as e:
        print(f"[TOP REJECTION REASONS ERROR] {e}")
        return jsonify([])


@app.route('/api/data/insured-claims-center', methods=['GET'])
def get_insured_claims_center():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        where_sql, params = _build_claim_filters(request.args)

        cursor.execute(
            f"""
            SELECT
                ISNULL(site_gestion_theo, 'N/A') AS centre,
                COUNT(DISTINCT id_sinistre) AS nb_insured_claims_treated,
                SUM(CASE WHEN date_cloture IS NOT NULL AND date_survenance IS NOT NULL AND DATEDIFF(DAY, date_survenance, date_cloture) <= 5 THEN 1 ELSE 0 END) AS nb_insured_claims_treated_5d,
                AVG(CASE WHEN date_cloture IS NOT NULL AND date_survenance IS NOT NULL THEN DATEDIFF(DAY, date_survenance, date_cloture) * 1.0 END) AS tat_5_days,
                CAST(5 AS DECIMAL(10,2)) AS target_5days,
                CAST(5.5 AS DECIMAL(10,2)) AS tolerance_5days
            FROM dbo.claims_closed
            {where_sql}
            GROUP BY site_gestion_theo
            ORDER BY nb_insured_claims_treated DESC, centre ASC
            """,
            params,
        )
        rows = _rows_to_dicts(cursor)
        conn.close()
        return jsonify(rows)
    except Exception as e:
        print(f"[INSURED CLAIMS CENTER ERROR] {e}")
        return jsonify([])


@app.route('/api/data/provider-claims-center', methods=['GET'])
def get_provider_claims_center():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        where_sql, params = _build_claim_filters(request.args)

        cursor.execute(
            f"""
            SELECT
                ISNULL(c.site_gestion_theo, 'N/A') AS centre,
                CAST(AVG(CASE WHEN c.date_cloture IS NOT NULL AND c.date_survenance IS NOT NULL AND DATEDIFF(DAY, c.date_survenance, c.date_cloture) <= 20 THEN 100.0 ELSE 0.0 END) AS DECIMAL(10,2)) AS provider_claim_20days,
                SUM(CASE WHEN c.date_cloture IS NOT NULL AND c.date_survenance IS NOT NULL AND DATEDIFF(DAY, c.date_survenance, c.date_cloture) <= 20 THEN 1 ELSE 0 END) AS nb_claim_treated_20d,
                SUM(CASE WHEN c.date_cloture IS NOT NULL AND c.date_survenance IS NOT NULL AND DATEDIFF(DAY, c.date_survenance, c.date_cloture) <= 30 THEN 1 ELSE 0 END) AS nb_claim_treated_30d,
                COUNT(DISTINCT c.id_sinistre) AS nb_claims_treated,
                COUNT(l.id_ligne) AS nb_ligne_claims_treated,
                COUNT(DISTINCT CASE WHEN l.id_ligne IS NOT NULL THEN c.id_sinistre END) AS nb_claims_ss,
                COUNT(l.id_ligne) AS nb_ligne_claims_ss
            FROM dbo.claims_closed c
            LEFT JOIN dbo.claims_closed_lines l
              ON l.id_sinistre = c.id_sinistre AND l.num_sinistre = c.num_sinistre
            {where_sql}
            GROUP BY c.site_gestion_theo
            ORDER BY nb_claims_treated DESC, centre ASC
            """,
            params,
        )
        rows = _rows_to_dicts(cursor)
        conn.close()
        return jsonify(rows)
    except Exception as e:
        print(f"[PROVIDER CLAIMS CENTER ERROR] {e}")
        return jsonify([])

print(app.url_map)

if __name__ == '__main__':
    print("STARTING FLASK SERVICE on 0.0.0.0:5001")
    app.run(debug=True, host='0.0.0.0', port=5001)